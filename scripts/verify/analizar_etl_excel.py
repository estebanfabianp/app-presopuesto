#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analizar el archivo ETL de cuenta bancaria
"""

import sys
from pathlib import Path
import openpyxl

ROOT_DIR = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT_DIR))

# Leer el Excel
excel_path = ROOT_DIR / "cuenta_bancaria.xlsx"

print("\n" + "=" * 100)
print(f"ANALIZANDO: {excel_path}")
print("=" * 100)

wb = openpyxl.load_workbook(excel_path)
print(f"\nHojas disponibles: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n" + "-" * 100)
    print(f"HOJA: {sheet_name}")
    print("-" * 100)
    
    # Obtener las 5 primeras filas para ver estructura
    print(f"\nPrimeras filas:")
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
        print(f"  Fila {i}: {row}")
    
    # Contar filas totales
    max_row = ws.max_row
    print(f"\nTotal de filas: {max_row}")
    
    # Si hay datos, mostrar últimas filas
    if max_row > 5:
        print(f"\nUltimas 5 filas:")
        for i, row in enumerate(ws.iter_rows(min_row=max_row-4, max_row=max_row, values_only=True), max_row-4):
            print(f"  Fila {i}: {row}")
    
    # Calcular suma si hay columna de montos
    print(f"\nAnalizando columnas para encontrar montos...")
    first_row = list(ws.iter_rows(max_row=1, values_only=True))[0]
    print(f"  Encabezados: {first_row}")
    
    # Buscar columna con montos
    for col_idx, header in enumerate(first_row):
        if header and ('monto' in str(header).lower() or 'valor' in str(header).lower() or 'amount' in str(header).lower()):
            print(f"\n  Columna '{header}' encontrada (posición {col_idx+1})")
            
            total = 0
            count = 0
            for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=col_idx+1, max_col=col_idx+1, values_only=True):
                if row[0] is not None:
                    try:
                        total += float(row[0])
                        count += 1
                    except:
                        pass
            
            print(f"    Total registros con monto: {count}")
            print(f"    Suma total: {total:.2f}")

wb.close()
